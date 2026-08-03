#!/usr/bin/env python3
"""驗收產出檔。

骨架列數改成跟著當年度來源走之後，拿舊年度的勾稽檔逐格比對已經沒有意義
（來源少一個科目，產出就少一列，跟參考檔從那一列起全部錯開，但那是對的）。
改成檢查產出自己該成立的性質——這些性質不需要參考檔，而且比對格更嚴：
查表公式指到的那一列，科目名必須就是這一列自己的科目名。

用法：python3 test_build.py 產出.xlsx 12.平衡xxxx.xlsx 3.現流xxxx.xlsx
"""
import re
import sys

import openpyxl

import build_gouji as B

TOTAL_SHEET = B.TOTAL_SHEET
NOTE_SHEET = B.NOTE_SHEET

VLOOKUP_RE = re.compile(
    r"^=VLOOKUP\(A(\d+),'?" + re.escape(TOTAL_SHEET) + r"'?!\$A\$(\d+):\$[A-Z]+\$(\d+),\$K\$1,FALSE\)$")
OFFSET_RE = re.compile(
    r"^=OFFSET\('?" + re.escape(TOTAL_SHEET) + r"'?!\$A\$(\d+),0,\$K\$1-1\)$")
SUM3D_RE = re.compile(r"^=SUM\([^!]+!\$?[A-Z]{1,2}\$?(\d+)\)$")


def txt(v):
    return None if v is None else str(v).strip()


def label(v):
    """0／空字串／公式都不是科目名。"""
    s = txt(v)
    return None if s in (None, "", "0") or s.startswith("=") else s


def main():
    out_path = sys.argv[1]
    wb = openpyxl.load_workbook(out_path)
    wt = wb[TOTAL_SHEET]
    names = [n for n in wb.sheetnames if n not in (NOTE_SHEET, TOTAL_SHEET)]
    assert len(names) == B.N_UNI, f"明細表 {len(names)} 張，預期 {B.N_UNI} 張"

    blocks = B.scan_blocks(wt)
    skel_end = min(b.header for b in blocks.values()) - 2
    print(f"骨架 {B.SKEL_START}–{skel_end} 列；資料區塊 "
          + "、".join(f"{n} {b.start}–{b.end}" for n, b in blocks.items()))

    fails = []

    def check(cond, msg):
        if not cond:
            fails.append(msg)

    # ---- 1. 骨架的科目清單 = 來源的科目清單 ----
    if len(sys.argv) > 3:
        wsb = B.pick_sheet(openpyxl.load_workbook(sys.argv[2], data_only=True),
                           tag="1.本年預算數")
        _, rows = B.read_matrix(wsb)
        a, l = B.split_sections(rows, wsb.title)
        want_a = [txt(lb) for lb, _ in a + l
                  if lb is not None and B.indent(lb) <= B.MAX_LEVEL]
        _, rc = B.read_matrix(B.pick_sheet(
            openpyxl.load_workbook(sys.argv[3], data_only=True), detail=True))
        want_h = [txt(lb) for lb, _ in rc if lb is not None]
        # 比對用多重集合不用序列：骨架的排序是人工調過的（「什項負債」被排到
        # 「遞延負債」前面，它下面整組調整列才不會被拆散），照來源順序重排是錯的。
        for side, col, want in (("A 欄科目", 1, want_a), ("H 欄現流項目", 8, want_h)):
            got = [x for x in (label(wt.cell(r, col).value)
                               for r in range(B.SKEL_START + B.DETAIL_SHIFT, skel_end + 1)) if x]
            check(sorted(got) == sorted(want),
                  f"{side}與來源不符（產出 {len(got)} 項、來源 {len(want)} 項；"
                  f"多 {sorted(set(got) - set(want))}、缺 {sorted(set(want) - set(got))}）")
            if got != want:
                print(f"  （{side}與來源同一批、順序不同：骨架的人工排序優先）")
        print(f"✓ 骨架科目清單與來源一致（平衡表 {len(want_a)} 項、現流 {len(want_h)} 項）")
    else:
        print("· 略過來源比對（未指定來源檔）")

    # ---- 2. 查表公式指到的列，名稱必須就是這一列自己的名稱 ----
    n_off = n_look = 0
    for name in names:
        ws = wb[name]
        for r in range(B.SKEL_START, skel_end + 1):
            j = ws.cell(r, 10).value
            m = OFFSET_RE.match(str(j))
            if m:
                n_off += 1
                check(txt(wt.cell(int(m.group(1)), 1).value) == label(ws.cell(r, 8).value),
                      f"{name} J{r} 指到總表第 {m.group(1)} 列 "
                      f"{txt(wt.cell(int(m.group(1)), 1).value)!r}，"
                      f"但本列項目是 {label(ws.cell(r, 8).value)!r}")
            for c in (2, 12):
                m = VLOOKUP_RE.match(str(ws.cell(r, c).value))
                if not m:
                    continue
                n_look += 1
                key = label(ws.cell(int(m.group(1)), 1).value)
                lo, hi = int(m.group(2)), int(m.group(3))
                check(any(txt(wt.cell(br, 1).value) == key for br in range(lo, hi + 1)),
                      f"{name} {'B' if c == 2 else 'L'}{r} 查 {key!r}，"
                      f"總表 {lo}–{hi} 列查不到")
    print(f"· 查表公式 {n_off} 條 OFFSET、{n_look} 條 VLOOKUP")

    # ---- 3. 明細表與總表逐列對齊；跨表加總指同一列 ----
    wording = set()
    for name in names:
        ws = wb[name]
        for r in range(2, skel_end + 1):
            for c in (1, 8):
                a_, b_ = label(ws.cell(r, c).value), label(wt.cell(r, c).value)
                check((a_ is None) == (b_ is None),
                      f"未對齊 {name} {'AH'[c // 8]}{r}: {a_!r} vs 總表 {b_!r}")
                if a_ != b_ and a_ is not None and b_ is not None:
                    wording.add(f"{'AH'[c // 8]}{r}")
    if wording:
        print(f"  （列有對齊、只是用字略有出入：{'、'.join(sorted(wording))}）")
    for r in range(1, skel_end + 1):
        for c in range(1, 20):
            m = SUM3D_RE.match(str(wt.cell(r, c).value))
            check(not m or int(m.group(1)) == r,
                  f"總表 {r},{c} 跨表加總指到第 {m.group(1) if m else '?'} 列")

    for f in fails[:25]:
        print("  ✗ " + f)
    if len(fails) > 25:
        print(f"  …另有 {len(fails) - 25} 項")
    assert not fails, f"{len(fails)} 項檢查未通過"
    print("✓ 47 張明細表與總表逐列對齊，總表跨表加總全部指同一列")


if __name__ == "__main__":
    main()
