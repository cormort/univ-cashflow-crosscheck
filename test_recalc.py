#!/usr/bin/env python3
"""用 LibreOffice 真的重算產出檔，再檢查勾稽公式算出來對不對。

`test_build.py` 只讀公式字串，抓不到「公式長得對、算起來錯」的問題——
總表的骨架比明細樣板低一列，曾經因此整張表的參照多位移一列，靜態檢查
兩邊都是綠的，重算才噴 #VALUE!。這支就是補那個缺口。

檢查四項（前三項不過就 exit 1）：
  1. 重算後沒有任何錯誤值（#REF!/#N/A/#VALUE!/…）
  2. 四個檢查欄逐格算對（47 張表）：M = L-G、O = L-B、P = J-K、Q = D-F
  3. 總表 T/U 欄列出的「不一致學校」與逐張表清點的結果相符
  4. 底稿自帶的平衡列（資產合計 - 負債 - 淨值）：
     B 欄（上年，純從來源查表）必須 47 校全為 0——這是資料有沒有正確搬進去的鐵證；
     G 欄（本年 = 上年 + 調整）只報數不判定，調整分錄本來就要人工填。

用法：python3 test_recalc.py 大學116現金流量表勾稽檔-47所大學.xlsx
沒裝 LibreOffice 就整支跳過（exit 0）；安裝：brew install --cask libreoffice
"""
import pathlib
import re
import shutil
import subprocess
import sys
import tempfile

import openpyxl
from openpyxl.utils import get_column_letter

TOTAL_SHEET = "大學-自己調整"
NOTE_SHEET = "筆記"
SKEL_START = 7          # 產出檔骨架的第一列
ERRORS = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!")
BALANCE_RE = re.compile(r"^=([BG])(\d+)-\1(\d+)-\1(\d+)$")
SOFFICE_PATHS = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
    "/usr/bin/libreoffice",
]


def find_soffice():
    return next((p for p in SOFFICE_PATHS if pathlib.Path(p).exists()),
                shutil.which("soffice") or shutil.which("libreoffice"))


def recalc(soffice, src, workdir):
    """轉一次檔 LibreOffice 就會重算，並把結果寫進快取值。"""
    src = pathlib.Path(src)
    copy = pathlib.Path(workdir) / src.name        # 別讓它就地覆寫原檔
    shutil.copy(src, copy)
    out = pathlib.Path(workdir) / "out"
    subprocess.run(
        [soffice, f"-env:UserInstallation=file://{workdir}/profile",
         "--headless", "--norestore", "--convert-to", "xlsx",
         "--outdir", str(out), str(copy)],
        check=True, capture_output=True, timeout=900)
    return out / src.name


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else 0


def check(path, soffice):
    """回傳未通過的項目清單（空的代表全過）。"""
    with tempfile.TemporaryDirectory() as td:
        done = recalc(soffice, path, td)
        val = openpyxl.load_workbook(done, data_only=True)
        fml = openpyxl.load_workbook(done)

        names = [n for n in val.sheetnames if n not in (NOTE_SHEET, TOTAL_SHEET)]
        wt = val[TOTAL_SHEET]
        skel_end = next(r for r in range(10, wt.max_row + 1)
                        if str(wt.cell(r, 1).value).strip()
                        in ("科目", "項      目")) - 2
        fails = []

        # 1) 錯誤值
        errs = [f"{n}!{c.coordinate}={c.value}"
                for n in val.sheetnames for row in val[n].iter_rows() for c in row
                if any(e in str(c.value) for e in ERRORS)]
        if errs:
            fails.append(f"重算後有 {len(errs)} 格錯誤值，例：{errs[:5]}")
        print(f"  重算後錯誤值：{len(errs)} 格" + ("" if errs else "　✓"))

        # 2/3) 檢查欄算出來的值
        # 四個檢查欄都要驗：build_gouji 逐列重新生成它們，這裡確認真的算出那個值
        for col, (x, y), name in ((13, (12, 7), "M=L-G"), (15, (12, 2), "O=L-B"),
                                  (16, (10, 11), "P=J-K"), (17, (4, 6), "Q=D-F")):
            bad = []
            for n in names:
                ws, wf = val[n], fml[n]
                for r in range(6, skel_end + 1):
                    if not str(wf.cell(r, col).value or "").startswith("="):
                        continue
                    got = num(ws.cell(r, col).value)
                    want = num(ws.cell(r, x).value) - num(ws.cell(r, y).value)
                    if abs(got - want) > 0.5:
                        bad.append(f"{n}!{chr(64 + col)}{r} 算出 {got} 應為 {want}")
            if bad:
                fails.append(f"{name} 有 {len(bad)} 格不符，例：{bad[:3]}")
            print(f"  {name}（47 張表）：{len(bad)} 格不符" + ("" if bad else "　✓"))

        # 3) 總表的不一致學校清單：對照自己逐張表清點的結果
        wrong = []
        for col, src in ((20, 13), (21, 16)):
            for r in range(SKEL_START, skel_end + 1):
                if not str(fml[TOTAL_SHEET].cell(r, col).value or "").startswith("="):
                    continue
                got = sorted(str(val[TOTAL_SHEET].cell(r, col).value or "").split())
                want = sorted(n for n in names if abs(num(val[n].cell(r, src).value)) > 0.5)
                if got != want:
                    wrong.append(f"{get_column_letter(col)}{r} 列出 {got} 應為 {want}")
        if wrong:
            fails.append(f"總表不一致學校清單有 {len(wrong)} 列不符，例：{wrong[:3]}")
        print(f"  總表不一致學校清單（T/U 欄）：{len(wrong)} 列不符"
              + ("" if wrong else "　✓"))

        # 4) 底稿自帶的平衡列
        spots = [(r, c) for r in range(6, skel_end + 1) for c in (2, 7)
                 if BALANCE_RE.match(str(fml[names[0]].cell(r, c).value or ""))]
        if not spots:
            fails.append("找不到平衡檢查列（=B資產合計-B負債-B淨值）")
        for r, c in spots:
            off = [(n, val[n].cell(r, c).value or 0) for n in names]
            off = [(n, v) for n, v in off if abs(num(v)) > 0.5]
            if c == 2:      # 上年：純從來源查表，不平衡就是搬運出錯
                if off:
                    fails.append(f"上年平衡列 B{r} 有 {len(off)} 校不為 0：{off[:3]}")
                print(f"  上年平衡 B{r}：{len(off)}/{len(names)} 校不為 0"
                      + ("" if off else "　✓"))
            else:           # 本年：調整分錄未填就會不為 0，只報數不判定
                print(f"  本年平衡 G{r}：{len(off)}/{len(names)} 校不為 0"
                      f"（調整分錄未填屬正常）" + (f"　例 {off[:3]}" if off else "　✓"))
        return fails


def main():
    soffice = find_soffice()
    if not soffice:
        print("· 略過重算檢查：找不到 LibreOffice"
              "（brew install --cask libreoffice）")
        return
    bad = {}
    for path in sys.argv[1:]:
        print(f"\n{pathlib.Path(path).name}")
        fails = check(path, soffice)
        if fails:
            bad[path] = fails
    for path, fails in bad.items():
        print(f"\n✗ {pathlib.Path(path).name}")
        for f in fails:
            print(f"    {f}")
    if bad:
        raise SystemExit(1)
    print("\n✓ 重算無錯誤值、檢查欄算得對、上年平衡列 47 校全為 0")


if __name__ == "__main__":
    main()
