#!/usr/bin/env python3
"""從別的勾稽表抽出「平衡表科目調整 ↔ 現流項目」配對，評估能不能併進主表。

只讀不改：不動來源檔、不動 勾稽配對表.json。要併是另一個決定。

配對藏在兩個地方，有一個就抽得出來：
  1. 借貸金額是公式、直接指到現流列（`=J50*-1`）——我們這邊約佔八成
  2. C/E 的借貸代號與 I 欄的代號相同者互相配對——約佔兩成
第 2 種只要求「該檔案內部代號自洽」，不需要跟我們的編碼一致。

用法：python3 extract_pairs.py 別人的勾稽表.xlsx [輸出.csv]
他機關多半還是 .xls（BIFF8），丟進來會自動用 LibreOffice 轉一份暫存檔。
"""
import collections
import csv
import json
import pathlib
import re
import shutil
import subprocess
import sys
import tempfile

import openpyxl

SOFFICE = ("soffice", "/Applications/LibreOffice.app/Contents/MacOS/soffice")


def as_xlsx(path):
    """openpyxl 讀不了 BIFF8，先轉檔。回傳可讀的路徑（.xlsx 原樣奉還）。"""
    if not str(path).lower().endswith(".xls"):
        return path
    exe = next((x for x in SOFFICE if shutil.which(x) or pathlib.Path(x).exists()), None)
    if exe is None:
        raise SystemExit("✗ 這是 .xls，需要 LibreOffice 轉檔：brew install --cask libreoffice")
    tmp = tempfile.mkdtemp(prefix="gouji-xls-")
    subprocess.run([exe, "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
                   check=True, capture_output=True)
    out = pathlib.Path(tmp) / (pathlib.Path(path).stem + ".xlsx")
    if not out.exists():
        raise SystemExit(f"✗ 轉檔失敗，LibreOffice 沒有產出 {out.name}")
    print(f"（.xls 已轉為暫存 xlsx：{out}）")
    return out

MASTER_FILE = "現流代號主檔.json"
PAIRING_FILE = "勾稽配對表.json"
HDR = re.compile(r"業務活動之現金流量")
COL = {"科目": 1, "借代號": 3, "借金額": 4, "貸代號": 5, "貸金額": 6}


def locate(ws):
    """右半邊現流區塊在第幾欄——不是每份都在 H。

    左半邊的調整欄 C/D/E/F 各家一致，右半邊會被插進去的欄推著走：
    醫療藥品多一欄「平衡表數」推到 I、高中多「平衡表數＋檢誤」推到 J〔實測〕。
    認第一節的節名定位，代號與金額固定接在右邊兩欄。
    """
    for r in range(1, 20):
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if isinstance(v, str) and HDR.search(v):
                return {**COL, "現流項目": c, "現流代號": c + 1, "現流數": c + 2}
    raise SystemExit("✗ 找不到「業務活動之現金流量」，版面與預期不符")


def indent(s):
    n = 0
    for ch in str(s):
        if ch != "　":
            break
        n += 1
    return n


def skeleton(ws, COL):
    """骨架範圍：從 H 欄開始有項目，到骨架下方的資料區開始為止。

    骨架下方是什麼隨檔案而異——完成檔是鏡射公式（A 欄 `='大學-自己調整'!A139`）、
    總表是資料區塊表頭（A 欄 `科目`）。兩種都要擋，否則會把資料區的數字
    當成借貸代號抽進來。
    """
    lo = next((r for r in range(1, 40) if ws.cell(r, COL["現流項目"]).value not in (None, "")), None)
    if lo is None:
        raise SystemExit(f"✗ 這份檔案的第 {COL['現流項目']} 欄沒有現流項目，版面與預期不符")
    hi = ws.max_row
    for r in range(lo + 5, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and (a.startswith("=") or a.strip() in ("科目", "項      目")):
            hi = r - 1
            break
    while hi > lo and all(ws.cell(hi, c).value in (None, "")
                          for c in (1, COL["現流項目"])):
        hi -= 1
    return lo, hi


def extract(path):
    wb = openpyxl.load_workbook(as_xlsx(path))
    ws = wb[wb.sheetnames[0]] if len(wb.sheetnames) < 3 else wb[wb.sheetnames[2]]
    COL = locate(ws)
    # 借貸金額指回現流數那一欄（我們的檔是 J，別人的可能是 K/L）
    jref = re.compile(r"(?<![A-Z$])%s(\d+)"
                      % openpyxl.utils.get_column_letter(COL["現流數"]))
    lo, hi = skeleton(ws, COL)

    stack, item, parent = [], {}, {}
    for r in range(lo, hi + 1):
        h = ws.cell(r, COL["現流項目"]).value
        if h in (None, ""):
            continue
        d = indent(h)
        del stack[d:]
        while len(stack) < d:
            stack.append("")
        stack.append(str(h).strip())
        item[r] = stack[-1]
        parent[r] = next((x for x in reversed(stack[:-1]) if x), "")
    by_code = {str(ws.cell(r, COL["現流代號"]).value).strip(): r for r in range(lo, hi + 1)
               if ws.cell(r, COL["現流代號"]).value not in (None, "")}

    def owner(r):
        for x in range(r, lo - 1, -1):
            v = ws.cell(x, COL["科目"]).value
            if isinstance(v, str) and v.strip() not in ("", "0"):
                return v.strip()
        return "?"

    pairs, seen, how = [], collections.Counter(), collections.Counter()
    for r in range(lo, hi + 1):
        for cc, ac, side in ((COL["借代號"], COL["借金額"], "借"),
                             (COL["貸代號"], COL["貸金額"], "貸")):
            v = ws.cell(r, cc).value
            if v in (None, "") or (isinstance(v, str) and v.startswith("=")):
                continue
            amt = str(ws.cell(r, ac).value or "")
            j = {int(m.group(1)) for m in jref.finditer(amt)}
            by_f = next(iter(j)) if len(j) == 1 else None
            tgt = by_f if by_f in item else by_code.get(str(v).strip())
            how["公式" if by_f in item else ("代號" if tgt else "無法解析")] += 1
            o = owner(r)
            seen[(o, side)] += 1
            pairs.append({"平衡表科目": o, "借貸": side, "第N筆": seen[(o, side)],
                          "原代號": str(v).strip(),
                          "現流項目": item.get(tgt, ""), "現流上層": parent.get(tgt, ""),
                          "金額公式": amt, "來源": pathlib.Path(path).name})
    return ws.title, (lo, hi), pairs, how


def main():
    path = sys.argv[1]
    title, (lo, hi), pairs, how = extract(path)
    print(f"工作表「{title}」骨架 {lo}–{hi} 列；抽出 {len(pairs)} 筆配對　來源 {dict(how)}\n")

    if not pathlib.Path(MASTER_FILE).exists():
        print(f"（找不到 {MASTER_FILE}，跳過與公版的比對）")
    else:
        master = json.loads(pathlib.Path(MASTER_FILE).read_text())["代號"]
        coded = {k.split("\t")[1] for k in master}
        # 公版裡的群組不編碼，配對指到群組是資料怪異，不是名稱對不上
        chart = coded | {k.split("\t")[0] for k in master if k.split("\t")[0]}
        got = {p["現流項目"] for p in pairs if p["現流項目"]}
        hit, group, alien = got & coded, (got & chart) - coded, got - chart
        verdict = "　→ 抽不到現流項目，先確認版面" if not got else (
            "　→ 可直接併" if not alien else "　→ 需要名稱對照")
        print(f"① 名稱與公版的重疊：{len(got & chart)}/{len(got)}"
              f"（{100 * len(got & chart) / max(len(got), 1):.0f}%）" + verdict)
        for n in sorted(alien)[:8]:
            print(f"      公版沒有：{n[:40]}")
        for n in sorted(group)[:5]:
            print(f"      指到群組（公版不編碼，請確認）：{n[:34]}")
        bad = sum(1 for p in pairs if not p["現流項目"])
        if bad:
            print(f"      另有 {bad} 筆對不到現流項目"
                  f"（可能是該檔的孤兒代號，或版面與預期不符）")

        have = set(json.loads(pathlib.Path(PAIRING_FILE).read_text())["配對"].values()) \
            if pathlib.Path(PAIRING_FILE).exists() else set()
        have_names = {k.split("\t")[1] for k in have}
        new = hit - have_names
        print(f"\n② 能補的缺口：{len(new)} 個現流項目是我們沒有配對過的"
              f"（公版可編碼 {len(master)} 個、目前覆蓋 {len(have)} 個）")
        for n in sorted(new)[:10]:
            print(f"      {n[:44]}")

    out = sys.argv[2] if len(sys.argv) > 2 else None
    if out:
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(pairs[0]))
            w.writeheader()
            w.writerows(pairs)
        print(f"\n→ {out}")
    print("\n（只讀不改；要併進 勾稽配對表.json 是另一個決定，"
          "建議先跑歸零檢定確認配對正確）")


if __name__ == "__main__":
    main()
