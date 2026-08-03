#!/usr/bin/env python3
"""由「12.平衡」與「3.現流」兩份來源檔，產生大學現金流量表勾稽底稿。

作法：以既有勾稽檔為範本（保留全部公式與格式），
      1) 47 張明細表逐格取眾數，消掉各表之間的漂移
      2) 骨架列數跟著當年度的科目／項目清單走：來源沒有的科目刪列、
         多出來的補列，範本累積的借貸代號與調整公式依名稱重新錨定
      3) 五個原始資料區塊只寫一份在總表底下
      4) 明細表的 VLOOKUP 範圍 / OFFSET 錨點改指總表，並依新列數重算
"""
import argparse
import collections
import difflib
import json
import pathlib
import re
import sys

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


class BuildError(Exception):
    """來源檔或範本不符預期，訊息直接給使用者看。"""

TOTAL_SHEET = "大學-自己調整"
NOTE_SHEET = "筆記"
SKEL_ROWS = 135          # 範本明細表的骨架列數（產出的骨架列數依來源而定）
SKEL_START = 6           # 骨架第一列（A6=資產、H6=▼業務活動之現金流量）
DETAIL_SHIFT = 1         # 明細表整體下移，讓列號與總表對齊
FROZEN_ROWS = 3          # 前三列（標題）不下移
SKEL_COLS = 18           # A~R
N_UNI = 47
TOTAL_COL = 2            # 資料區塊的「合計」欄
BLOCK_LAST_COL = TOTAL_COL + N_UNI   # 49 = AW

# 骨架一列同時放兩份互不相干的清單：A~G 是平衡表科目、H~ 是現流項目。
# 公式裡的列參照要查哪一側的對照表，由「被參照的欄」決定，不是由公式所在的欄。
H_SIDE_COLS = frozenset((8, 9, 10, 11, 16))   # H 項目、I 代號、J 現流數、K 反算、P 差異
MAX_LEVEL = 2            # 平衡表科目取到第幾層（縮排的全形空白數）

# 五個資料區塊的順序（範本與產出一致）
BLOCK_NAMES = ["本年資產", "本年負債淨值", "上年資產", "上年負債淨值", "現流"]


# --------------------------------------------------------------------------
# 來源檔讀取
# --------------------------------------------------------------------------
def has_unis(ws, header_row=4):
    """第 4 列 C~AW 有 47 個校名 = 這是各校明細表，不是彙總表或空白表。"""
    return all(ws.cell(header_row, c).value not in (None, "")
               for c in range(TOTAL_COL + 1, BLOCK_LAST_COL + 1))


def pick_sheet(wb, tag=None, detail=False):
    """依 A1 標記或「有 47 校」的結構挑工作表。

    表名逐年不同（`細修0814`／`細`／`明細`／`Sheet1`），所以只認結構不認名字。
    排除「差異」表；優先取名稱帶最新四位日期後綴者，同日期取名稱最短者
    （避開 `114調後2`、`細修2` 這種暫存版）。
    """
    cands = []
    for name in wb.sheetnames:
        if "差異" in name:
            continue
        if tag is not None and str(wb[name]["A1"].value).strip() != tag:
            continue
        if detail and not has_unis(wb[name]):
            continue
        m = re.search(r"(\d{4})$", name)
        cands.append(((int(m.group(1)) if m else -1, -len(name)), name))
    if not cands:
        raise BuildError(f"找不到工作表（tag={tag!r} detail={detail}）")
    return wb[max(cands)[1]]


def read_matrix(ws, header_row=4):
    """回傳 (校名清單, [(科目, [合計, *47校])])，保留空白列以便切段。"""
    unis = [ws.cell(header_row, c).value for c in range(TOTAL_COL + 1, BLOCK_LAST_COL + 1)]
    if any(u is None for u in unis):
        raise BuildError(f"{ws.title}: 第 {header_row} 列的校名不足 {N_UNI} 校")
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        vals = [ws.cell(r, c).value for c in range(TOTAL_COL, BLOCK_LAST_COL + 1)]
        rows.append((label, vals))
    while rows and rows[-1][0] is None:
        rows.pop()
    return unis, rows


def split_sections(rows, sheet_title):
    """把平衡表切成 (資產段, 負債淨值段)，各自含最後的「合計」列。"""
    idx_total = [i for i, (lb, _) in enumerate(rows)
                 if isinstance(lb, str) and lb.split() == ["合", "計"]]
    idx_liab = [i for i, (lb, _) in enumerate(rows)
                if isinstance(lb, str) and lb.strip() == "負債"]
    if len(idx_total) < 2 or not idx_liab:
        raise BuildError(f"{sheet_title}: 找不到資產/負債的分段點（合計列 {len(idx_total)} 個）")
    return rows[: idx_total[0] + 1], rows[idx_liab[0]: idx_total[-1] + 1]


def parse_year(*sheets):
    """從來源工作表的標題解析本年度（民國年）。"""
    found = []
    for ws in sheets:
        for r in range(1, 8):
            for c in range(1, 8):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    m = re.search(r"中華民國(\d{2,3})年", v)
                    if m:
                        found.append((ws.title, v.strip(), int(m.group(1))))
                        break
            if found and found[-1][0] == ws.title:
                break
    if not found:
        raise BuildError("無法從來源檔標題解析年度，請用 --year 指定")
    years = {y for _, _, y in found}
    if len(years) > 1:
        print(f"! 兩份來源檔的年度不一致：{found}", file=sys.stderr)
    return found, min(years)


# --------------------------------------------------------------------------
# 範本：區塊掃描與眾數樣板
# --------------------------------------------------------------------------
Block = collections.namedtuple("Block", "header start end")


def scan_blocks(ws):
    """掃出總表在骨架之後的五個資料區塊（ws 需為 data_only 讀取）。"""
    headers = [r for r in range(SKEL_ROWS + 1, ws.max_row + 1)
               if isinstance(ws.cell(r, 1).value, str)
               and ws.cell(r, 1).value.strip() in ("科目", "項      目")]
    if len(headers) != len(BLOCK_NAMES):
        raise BuildError(f"{ws.title}: 掃到 {len(headers)} 個資料區塊，預期 {len(BLOCK_NAMES)} 個")
    blocks = {}
    for i, h in enumerate(headers):
        # 下一個區塊的表頭列前面是它的標示列（如「115(本年度)」），不屬於本區塊
        limit = headers[i + 1] - 1 if i + 1 < len(headers) else ws.max_row + 1
        end = h
        for r in range(h + 1, limit):
            if isinstance(ws.cell(r, 1).value, str) and ws.cell(r, 1).value.strip():
                end = r
        blocks[BLOCK_NAMES[i]] = Block(h, h + 1, end)
    return blocks


MIRROR_RE = re.compile(r"^='?" + re.escape(TOTAL_SHEET) + r"'?!\$?A\$?(\d+)$")


def scan_mirror(ws):
    """明細表骨架之後是總表資料區塊的鏡射公式；回傳 {明細列: 總表列}。"""
    mirror = {}
    for r in range(SKEL_ROWS + 1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str):
            m = MIRROR_RE.match(v.strip())
            if m:
                mirror[r] = int(m.group(1))
    if len(mirror) < 100:
        raise BuildError(f"{ws.title}: 只認出 {len(mirror)} 列鏡射公式，範本結構與預期不符")
    return mirror


def load_mirror(template_path, ws):
    """先讀鏡射公式；瘦身過的範本沒有那些列，改讀旁邊的 .mirror.json。"""
    try:
        return scan_mirror(ws)
    except BuildError:
        side = pathlib.Path(str(template_path) + ".mirror.json")
        if not side.exists():
            raise BuildError(f"範本沒有鏡射列，也找不到 {side.name}")
        return {int(k): v for k, v in json.loads(side.read_text()).items()}


def strip_template(src, dst):
    """把 14MB 範本瘦身成骨架範本：

    - 刪掉 47 張明細表的鏡射資料列（鏡射對照另存 JSON）
    - 清掉總表資料區塊裡的金額，只留科目名稱
      （產生時本來就只讀科目名稱來對列，金額一律重寫；
       範本因此不帶任何一年的預算數字，可以安心公開）
    """
    wb = openpyxl.load_workbook(src)
    names = [n for n in wb.sheetnames if n not in (NOTE_SHEET, TOTAL_SHEET)]
    mirror = scan_mirror(wb[names[0]])
    for n in names:
        ws = wb[n]
        if ws.max_row > SKEL_ROWS:
            ws.delete_rows(SKEL_ROWS + 1, ws.max_row - SKEL_ROWS + 1)
    wt = wb[TOTAL_SHEET]
    wiped = 0
    for r in range(SKEL_ROWS + 1, wt.max_row + 1):
        for c in range(TOTAL_COL, BLOCK_LAST_COL + 1):
            if wt.cell(r, c).value is not None:
                put(wt, r, c, None)
                wiped += 1

    # 47 張明細表一律換成眾數樣板：抹掉各校去年的人工填數與表名註記，
    # 範本因此不含任何一年的金額。（產生時本來就會取眾數，結果不變）
    tpl, _ = modal_template(wb, names)
    plugs = 0
    for i, n in enumerate(names, start=1):
        ws = wb[n]
        for (r, c), v in tpl.items():
            if ws.cell(r, c).value != v:
                if isinstance(ws.cell(r, c).value, (int, float)):
                    plugs += 1
                put(ws, r, c, v)
        put(ws, 1, 12, i)
        ws.title = n.split("-")[0]
    print(f"  已清掉範本裡 {wiped} 格舊年度金額、{plugs} 筆各校人工填數，表名註記也一併移除")
    wb.save(dst)
    pathlib.Path(str(dst) + ".mirror.json").write_text(
        json.dumps({str(k): v for k, v in mirror.items()}))
    print(f"✓ {dst}（{pathlib.Path(dst).stat().st_size / 1e6:.1f}MB）"
          f"＋ {pathlib.Path(dst).name}.mirror.json")


def modal_template(wb, detail_names):
    """47 張明細表逐格取眾數，回傳 (樣板, 分歧清單)。"""
    sheets = [wb[n] for n in detail_names]
    tpl, drift = {}, []
    for r in range(1, SKEL_ROWS + 1):
        for c in range(1, SKEL_COLS + 1):
            vals = [s.cell(r, c).value for s in sheets]
            cnt = collections.Counter(repr(v) for v in vals)
            best, n = cnt.most_common(1)[0]
            tpl[(r, c)] = vals[[repr(v) for v in vals].index(best)]
            if len(cnt) > 1:
                drift.append((r, c, cnt.most_common(3)))
    return tpl, drift


# --------------------------------------------------------------------------
# 骨架重排：讓骨架列數跟著當年度的科目／項目清單走
# --------------------------------------------------------------------------
Placed = collections.namedtuple("Placed", "row key label old head")


def indent(label):
    """科目名稱前的全形空白數 = 階層。"""
    n = 0
    for ch in str(label):
        if ch != "　":
            break
        n += 1
    return n


def parented(labels):
    """回傳每一項的顯示名稱，同名項目附上父項才看得出是哪一個。"""
    stack, out = [], []
    for lb in labels:
        depth = indent(lb)
        del stack[depth:]
        while len(stack) < depth:
            stack.append("")             # 來源偶有跳階，補空的佔位
        stack.append(str(lb).strip())
        parent = next((p for p in reversed(stack[:-1]) if p), None)
        out.append(f"{parent} › {stack[-1]}" if parent else stack[-1])
    return out


def skel_units(tpl, col, group_orphans):
    """把骨架切成單元，回傳 [(標籤, [列號…])]。

    平衡表側一個科目列會帶著它下面沒有科目名的調整列一起搬（G 欄的
    SUM(D29:D32) 就是在加這一組）；現流側一項一列。
    """
    units = []
    for r in range(SKEL_START, SKEL_ROWS + 1):
        v = tpl.get((r, col))
        label = v if isinstance(v, str) and v.strip() not in ("", "0") else None
        if label is not None:
            units.append((label, [r]))
        elif group_orphans and units:
            units[-1][1].append(r)
    return units


def replan(units, src_labels):
    """骨架單元 × 來源清單 → (新順序 [(來源序號, 來源原字串, 單元 or None)], 被移除的單元)。

    配對分兩步：
      1) 序列對齊。同名項目很多——「機械及設備」在撥入明細(+)、撥出明細(-)、
         固定資產之增置底下各有一個，只靠名字會配到別的父項底下去。
      2) 對齊時落單的單元再按名字補配。人工把「什項負債」排到「遞延負債」
         前面是有意的，序列對齊會把這種對調判成一刪一增，連同它底下整組
         調整列一起丟掉。
    排列照骨架的順序（同上，人工排序要留著），來源多出來的才插進去。
    名稱採來源的原字串，VLOOKUP 要跟資料區塊的鍵值逐字元相符。
    """
    skel_names = [str(lb).strip() for lb, _ in units]
    src_names = [str(lb).strip() for lb in src_labels]

    pair = {}
    for tag, i1, i2, j1, _ in difflib.SequenceMatcher(
            None, skel_names, src_names, autojunk=False).get_opcodes():
        if tag == "equal":
            for k in range(i2 - i1):
                pair[i1 + k] = j1 + k
    free = [j for j in range(len(src_names)) if j not in set(pair.values())]
    for i in range(len(units)):
        if i in pair:
            continue
        hit = next((j for j in free if src_names[j] == skel_names[i]), None)
        if hit is not None:
            pair[i] = hit
            free.remove(hit)

    plan = [(pair[i], src_labels[pair[i]], units[i]) for i in range(len(units)) if i in pair]
    for j in free:
        at = {jj: n for n, (jj, _, _) in enumerate(plan)}  # ponytail: 清單才幾十到幾百項，重建比維護索引省事
        prev = next((p for p in range(j - 1, -1, -1) if p in at), None)
        plan.insert(at[prev] + 1 if prev is not None else 0, (j, src_labels[j], None))
    return plan, [i for i in range(len(units)) if i not in pair]


def place(plan, start=SKEL_START):
    """指派新列號，回傳 (逐列紀錄, {舊列: 新列}, 骨架結束列)。"""
    rows, rowmap, r = [], {}, start
    for key, label, unit in plan:
        olds = unit[1] if unit else [None]
        for i, old in enumerate(olds):
            if old is not None:
                rowmap[old] = r
            rows.append(Placed(r, key, label, old, i == 0) if i == 0
                        else Placed(r, None, None, old, False))
            r += 1
    return rows, rowmap, r - 1


def fresh_balance(row, label, group_end, prev_block, cur_block, asset, prefix):
    """來源新增的平衡表科目：補機械公式，調整欄 C~F 留白給人工填。"""
    last = get_column_letter(BLOCK_LAST_COL)
    dr, cr = f"SUM(D{row}:D{group_end})", f"SUM(F{row}:F{group_end})"
    return {
        1: label,
        2: f"=VLOOKUP(A{row},{prefix}$A${prev_block.start}:${last}${prev_block.end},$K$1,FALSE)",
        7: f"=B{row}+{dr}-{cr}" if asset else f"=B{row}+{cr}-{dr}",
        12: f"=VLOOKUP(A{row},{prefix}$A${cur_block.start}:${last}${cur_block.end},$K$1,FALSE)",
        13: f"=L{row}-G{row}",
        15: f"=L{row}-B{row}",
        17: f"=D{row}-F{row}",
    }


def fresh_cashflow(row, label, block_row, prefix):
    """來源新增的現流項目：J 直接指到新區塊；K 的借貸代號沒有前例，留白。"""
    return {
        8: label,
        10: f"=OFFSET({prefix}$A${block_row},0,$K$1-1)",
        16: f"=J{row}-K{row}",
    }


# 檢查欄是純機械公式，一律指自己這一列
CHECK_COLS = {13: ("L", "G"), 15: ("L", "B"), 16: ("J", "K"), 17: ("D", "F")}
# 空白版要清掉的欄：C/E 借貸代號、D/F 調整金額、I 現流側代號——只有這五欄是人工判斷。
# K（由代號反算）不清：代號清空後它自然算 0，填代號時會自己亮起來，
# 而且裡面的 SUM 彙總編碼了現流的層級結構，手工重建很痛苦。
BLANK_COLS = (3, 4, 5, 6, 9)
VLOOKUP_KEY_RE = re.compile(r"^=VLOOKUP\(A(\d+),")


def canonical_check(col, row):
    """M=L-G、O=L-B、P=J-K、Q=D-F。

    範本這四欄被拖曳拖歪了 82 格（`O50=L50-B49`、`M133=L135-G137` 這種），
    照搬會讓人看到假差異，所以逐列重新生成而不沿用。
    """
    x, y = CHECK_COLS[col]
    return f"={x}{row}-{y}{row}"


SUMIF_C_RE = re.compile(r"\$C\$(\d+):\$C\$(\d+)")


def align_sumif(formula):
    """K 欄 SUMIF 的貸方(E/F)範圍要跟借方(C/D)一致。

    範本有 7 格把 E 範圍寫成 `$E$7:$E$135`（其餘 57 格都是 `$E$7:$E$127`）。
    SUMIF 的條件範圍比加總範圍大時，Excel 會自己把加總範圍延長補齊，
    等於多掃到底下那幾列欄位總計。
    """
    m = SUMIF_C_RE.search(formula or "")
    if not m:
        return formula
    lo, hi = m.group(1), m.group(2)
    return re.sub(r"\$([DEF])\$\d+:\$([DEF])\$\d+",
                  lambda x: f"${x.group(1)}${lo}:${x.group(2)}${hi}", formula)


def fix_lookup_key(formula, row, is_head):
    """B／L 的 VLOOKUP 一定是查「本列的科目」，範本有幾格被拖成查別列。"""
    m = VLOOKUP_KEY_RE.match(formula or "")
    if not m or int(m.group(1)) == row:
        return formula
    if not is_head:
        return None              # 調整列沒有科目可查，這格是拖曳拖出來的
    return VLOOKUP_KEY_RE.sub(f"=VLOOKUP(A{row},", formula, count=1)


def reposition(cells, map_a, map_h):
    """把 {(列, 欄): 值} 搬到新的骨架列號；表頭列不動，被刪掉的列丟棄。

    只搬內容不搬格式：openpyxl 沒有「連同樣式插入列」這回事。
    ponytail: 逐年插刪通常只有個位數列，錯位是外觀問題；真的礙眼再處理框線。
    """
    out, dropped = {}, []
    for (r, c), v in cells.items():
        if r < SKEL_START:
            out[(r, c)] = v
            continue
        nr = (map_h if c in H_SIDE_COLS else map_a).get(r)
        if nr is None:
            if v not in (None, ""):
                dropped.append((r, c, v))
        else:
            out[(nr, c)] = v
    return out, dropped


# --------------------------------------------------------------------------
# 公式改寫
# --------------------------------------------------------------------------
RANGE_RE = re.compile(r"(\$?)([A-Z]{1,2})(\$?)(\d+):(\$?)([A-Z]{1,2})(\$?)(\d+)")
CELL_RE = re.compile(r"(?<![\w$!:])(\$?)([A-Z]{1,2})(\$?)(\d+)(?![\w(:])")


def build_rowmap(old_blocks, new_blocks, old_ws, new_rows_by_block):
    """總表老列號 → 新列號，與 老列號 → 區塊名。

    用序列對齊（不是查表）比對新舊科目清單，因為現流表有重複的項目名稱
    （「收取利息」「機械及設備」都出現多次），純查表會對到第一筆。
    """
    rowmap, blockmap = {}, {}
    for name in BLOCK_NAMES:
        ob, nb = old_blocks[name], new_blocks[name]
        rowmap[ob.header] = nb.header
        for r in range(ob.header, ob.end + 1):
            blockmap[r] = name
        old_labels = [old_ws.cell(r, 1).value for r in range(ob.start, ob.end + 1)]
        new_labels = [lb for lb, _ in new_rows_by_block[name]]
        for tag, i1, i2, j1, _ in difflib.SequenceMatcher(
                None, old_labels, new_labels, autojunk=False).get_opcodes():
            if tag == "equal":
                for k in range(i2 - i1):
                    rowmap[ob.start + i1 + k] = nb.start + j1 + k
    return rowmap, blockmap


def rewrite(formula, skel, rowmap, blockmap, new_blocks, prefix):
    """把公式裡的列參照改到新位置。

    骨架列（<= SKEL_ROWS）依「被參照的欄」屬平衡表側還是現流側查對照表；
    更大的是資料區塊列，改指總表的新區塊（必要時加上總表前綴）。
    兩者一定要在同一輪換掉：骨架變高之後，重排後的骨架列號會跟舊的
    資料區塊列號重疊，分兩輪做會把已經換好的骨架列當成區塊列再換一次。
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula, True, set()

    map_a, map_h, shift, ref_shift = skel
    ok = True
    unresolved = set()

    def skel_row(col, row):
        """ref_shift：公式裡的列參照本來就帶著的位移。

        對照表是用明細樣板的座標建的。明細表的公式參照也是那個座標（ref_shift=0），
        但總表的骨架整體低一列，它的公式參照本來就是總表座標（ref_shift=1），
        要先換算回樣板座標再查表，否則整張總表會再多位移一列。
        """
        r = row - ref_shift
        if r < SKEL_START:
            return row + shift - ref_shift   # 第 4、5 列是表頭，不重排
        m = map_h if column_index_from_string(col) in H_SIDE_COLS else map_a
        new = m.get(r)
        return None if new is None else new + shift

    def is_skel(row):
        return row - ref_shift <= SKEL_ROWS

    def fix_range(m):
        nonlocal ok
        c1, r1, c2, r2 = m.group(2), int(m.group(4)), m.group(6), int(m.group(8))
        if is_skel(r1):
            n1, n2 = skel_row(c1, r1), skel_row(c2, r2)
            if n1 is None or n2 is None:
                ok = False
                unresolved.update({(c1, r1), (c2, r2)})
                return m.group(0)
            return (f"{m.group(1)}{c1}{m.group(3)}{n1}:"
                    f"{m.group(5)}{c2}{m.group(7)}{n2}")
        name = blockmap.get(r1)
        if name is None:
            ok = False
            unresolved.add((None, r1))
            return m.group(0)
        nb = new_blocks[name]
        return (f"{prefix}$A${nb.start}:${get_column_letter(BLOCK_LAST_COL)}${nb.end}")

    out = RANGE_RE.sub(fix_range, formula)

    def fix_cell(m):
        nonlocal ok
        col, row = m.group(2), int(m.group(4))
        if row <= FROZEN_ROWS:
            return m.group(0)                    # K1／L1 這種表頭參照不動
        if is_skel(row):
            new = skel_row(col, row)
            if new is None:
                ok = False
                unresolved.add((col, row))
                return m.group(0)
            return f"{m.group(1)}{col}{m.group(3)}{new}"
        new = rowmap.get(row)
        if new is None:
            ok = False
            unresolved.add((None, row))
            return m.group(0)
        return f"{prefix}${col}${new}"

    out = CELL_RE.sub(fix_cell, out)
    return out, ok, unresolved


# --------------------------------------------------------------------------
# 寫入
# --------------------------------------------------------------------------
def put(ws, r, c, value):
    """寫入儲存格；合併儲存格的非左上角略過（openpyxl 不允許寫入）。"""
    cell = ws.cell(r, c)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        return None
    cell.value = value
    return cell


def write_block(ws, marker_row, header_row, marker, head_label, unis, rows, num_fmt):
    put(ws, marker_row, 1, marker)
    put(ws, header_row, 1, head_label)
    put(ws, header_row, TOTAL_COL, "合    計")
    for i, u in enumerate(unis):
        put(ws, header_row, TOTAL_COL + 1 + i, u)
    for i, (label, vals) in enumerate(rows):
        r = header_row + 1 + i
        put(ws, r, 1, label)
        for j, v in enumerate(vals):
            cell = put(ws, r, TOTAL_COL + j, v)
            if cell is not None and isinstance(v, (int, float)):
                cell.number_format = num_fmt
    return header_row + len(rows)


DEFAULT_TEMPLATE = "template.xlsx"


def build(balance, cashflow, template=DEFAULT_TEMPLATE, output=None, year=None,
          bal_sheet=None, prev_sheet=None, cf_sheet=None, outdir=".", blank=False):
    """產生勾稽底稿，回傳 (輸出路徑, 報告文字)。"""
    lines = []
    def report(*a):
        lines.append(" ".join(str(x) for x in a))

    args = argparse.Namespace(balance=balance, cashflow=cashflow, template=template,
                              output=output, year=year, bal_sheet=bal_sheet,
                              prev_sheet=prev_sheet, cf_sheet=cf_sheet)

    # ---- 來源 ----
    wb_bal = openpyxl.load_workbook(args.balance, data_only=True)
    wb_cf = openpyxl.load_workbook(args.cashflow, data_only=True)

    ws_cur = wb_bal[args.bal_sheet] if args.bal_sheet else pick_sheet(wb_bal, tag="1.本年預算數")
    ws_prev = wb_bal[args.prev_sheet] if args.prev_sheet else pick_sheet(wb_bal, tag="5.上年調整後預算數")
    ws_cfd = wb_cf[args.cf_sheet] if args.cf_sheet else pick_sheet(wb_cf, detail=True)
    report(f"版本：{'空白版（調整欄留白待填）' if blank else '繼承版（沿用範本累積的調整對應）'}")
    report(f"來源工作表：本年={ws_cur.title}　上年={ws_prev.title}　現流={ws_cfd.title}\n")

    evidence, parsed = parse_year(ws_cur, ws_cfd)
    year = args.year or parsed
    report("年度解析：")
    for title, text, y in evidence:
        report(f"  {title}: {text} → {y}")
    report(f"  ⇒ 本年度 {year}、上年度 {year - 1}\n")

    unis_cur, rows_cur = read_matrix(ws_cur)
    unis_prev, rows_prev = read_matrix(ws_prev)
    unis_cf, rows_cf = read_matrix(ws_cfd)
    for tag, u in (("上年平衡表", unis_prev), ("現流表", unis_cf)):
        if u != unis_cur:
            raise BuildError(f"{tag}的校名順序與本年平衡表不一致")

    # 合計欄勾稽
    bad = []
    for tag, rows in (("本年", rows_cur), ("上年", rows_prev), ("現流", rows_cf)):
        for label, vals in rows:
            if not isinstance(vals[0], (int, float)):
                continue
            s = sum(v for v in vals[1:] if isinstance(v, (int, float)))
            if abs(s - vals[0]) > 0.5:
                bad.append((tag, label, vals[0], s))
    if bad:
        report("! 合計 ≠ 47 校橫加：")
        for t, lb, a, b in bad[:20]:
            report(f"   {t} {str(lb).strip()}: 合計 {a} vs 橫加 {b}")
        raise BuildError("來源資料未通過勾稽，中止")

    cur_a, cur_l = split_sections(rows_cur, ws_cur.title)
    prev_a, prev_l = split_sections(rows_prev, ws_prev.title)
    rows_cf = [r for r in rows_cf if r[0] is not None]
    new_data = {
        "本年資產": cur_a, "本年負債淨值": cur_l,
        "上年資產": prev_a, "上年負債淨值": prev_l,
        "現流": rows_cf,
    }
    markers = {
        "本年資產": f"{year}(本年度)", "本年負債淨值": f"{year}(本年度)",
        "上年資產": f"{year - 1}(上年度)", "上年負債淨值": f"{year - 1}(上年度)",
        "現流": "現金流量表",
    }

    # ---- 範本 ----
    wb = openpyxl.load_workbook(args.template)
    wbv = openpyxl.load_workbook(args.template, data_only=True)
    detail_names = [n for n in wb.sheetnames if n not in (NOTE_SHEET, TOTAL_SHEET)]
    if len(detail_names) != N_UNI:
        raise BuildError(f"範本有 {len(detail_names)} 張明細表，預期 {N_UNI} 張")

    old_total_blocks = scan_blocks(wbv[TOTAL_SHEET])
    mirror = load_mirror(args.template, wb[detail_names[0]])
    num_fmt = {name: wb[TOTAL_SHEET].cell(b.start, TOTAL_COL).number_format
               for name, b in old_total_blocks.items()}

    tpl, drift = modal_template(wb, detail_names)

    # ---- 骨架重排：列數跟著當年度的科目／項目清單走 ----
    bal_labels, bal_sections = [], []
    for sec, sec_rows in (("資產", cur_a), ("負債淨值", cur_l)):
        for lb, _ in sec_rows:
            if lb is not None and indent(lb) <= MAX_LEVEL:
                bal_labels.append(lb)
                bal_sections.append(sec)
    cf_labels = [lb for lb, _ in rows_cf]

    units_a, units_h = skel_units(tpl, 1, True), skel_units(tpl, 8, False)
    plan_a, gone_a = replan(units_a, bal_labels)
    plan_h, gone_h = replan(units_h, cf_labels)
    rows_a, map_a, end_a = place(plan_a)
    rows_h, map_h, end_h = place(plan_h)
    n_skel = max(end_a, end_h)

    # ---- 新區塊配置（全部寫在總表；總表的骨架比明細樣板低 DETAIL_SHIFT 列）----
    new_blocks, cursor = {}, n_skel + DETAIL_SHIFT + 1
    for name in BLOCK_NAMES:
        header = cursor + 1                       # cursor = 標示列
        new_blocks[name] = Block(header, header + 1, header + len(new_data[name]))
        cursor = new_blocks[name].end + 3
    rowmap_total, blockmap_total = build_rowmap(
        old_total_blocks, new_blocks, wbv[TOTAL_SHEET], new_data)
    cf_block_row = new_blocks["現流"].start      # 第 i 個現流項目就在區塊的第 i 列

    def fresh_cells(prefix, shift):
        """來源新增的科目／項目寫最終公式；沿用的列補上來源原字串的名稱。"""
        out = {}
        for p in rows_a:
            if not p.head:
                continue
            r = p.row + shift
            if p.old is not None:
                out[(r, 1)] = p.label
                continue
            asset = bal_sections[p.key] == "資產"
            out.update({(r, c): v for c, v in fresh_balance(
                r, p.label, r,
                new_blocks["上年資產" if asset else "上年負債淨值"],
                new_blocks["本年資產" if asset else "本年負債淨值"],
                asset, prefix).items()})
        for p in rows_h:
            if not p.head:
                continue
            r = p.row + shift
            cells = fresh_cashflow(r, p.label, cf_block_row + p.key, prefix)
            if p.old is not None:
                # 沿用的列只重寫名稱與 J：J 純粹是「指到本項目在區塊裡的那一列」，
                # 由名稱直接算比沿用舊列號可靠——範本的 J 是照舊區塊位置寫的，
                # 中間增刪科目就會指到別項去（K 的借貸代號才是人工資產，保留）。
                cells.pop(16, None)
            out.update({(r, c): v for c, v in cells.items()})
        return out
    # 明細表的區塊是總表的鏡射，經 mirror 轉一手即可
    rowmap_detail = {r: rowmap_total[t] for r, t in mirror.items() if t in rowmap_total}
    blockmap_detail = {r: blockmap_total[t] for r, t in mirror.items() if t in blockmap_total}

    # 產出檔的骨架只留有科目名的列，範本裡那些空列（連同指向它們的公式）會被丟掉
    void_rows = {c: {r for r in range(SKEL_START, SKEL_ROWS + 1) if r not in m}
                 for c, m in (("h", map_h), ("a", map_a))}
    head_a = {p.row for p in rows_a if p.head}

    def polish(orig, new, ok, col, row):
        """rewrite 之後的收尾。

        檢查欄一律重新生成——範本歪掉的本來就不該沿用，所以也不在意 rewrite
        有沒有對到位置；B／L 的 VLOOKUP 查找值拉回本列。
        """
        # 只清骨架本體：K1 是校序號、第 5 列是「借方／貸方」欄位標題，都得留著
        if blank and col in BLANK_COLS and row >= SKEL_START + DETAIL_SHIFT:
            return None, True
        if not isinstance(orig, str) or not orig.startswith("="):
            return new, ok
        if col in CHECK_COLS:
            return canonical_check(col, row), True
        if ok and col in (2, 12):
            return fix_lookup_key(new, row, row - DETAIL_SHIFT in head_a), True
        if ok and col == 11:
            return align_sumif(new), True
        return new, ok

    def why(unresolved, blockmap):
        """公式對不到新位置的原因，決定報告怎麼講。"""
        rows = {r for _, r in unresolved}
        if rows & set(blockmap):
            return "block"
        if all(c is not None and r in void_rows["h" if column_index_from_string(c)
                                                in H_SIDE_COLS else "a"]
               for c, r in unresolved if r <= SKEL_ROWS) and rows and max(rows) <= SKEL_ROWS:
            return "void"
        if any(r <= SKEL_ROWS for _, r in unresolved):
            return "removed"
        return "stale" if min(rows) > SKEL_ROWS else "lost"

    # ---- 明細表：套樣板 + 改寫公式 + 清掉原有資料區塊 ----
    tpl2, dropped_detail = reposition(tpl, map_a, map_h)
    fresh_detail = fresh_cells(f"'{TOTAL_SHEET}'!", DETAIL_SHIFT)
    failed = []
    renamed = []
    for idx, name in enumerate(detail_names, start=1):
        ws = wb[name]
        # openpyxl 的 insert_rows 不會跟著搬合併儲存格，得自己搬
        merges = [str(rng) for rng in ws.merged_cells.ranges]
        for rng in merges:
            ws.unmerge_cells(rng)
        ws.insert_rows(FROZEN_ROWS + 1, DETAIL_SHIFT)     # 空出與總表對齊用的列
        for rng in merges:
            mc = openpyxl.worksheet.cell_range.CellRange(rng)
            if mc.min_row > FROZEN_ROWS:
                mc.shift(row_shift=DETAIL_SHIFT)
            ws.merge_cells(str(mc))
        # 骨架區先清空：重排後某些列不再有對應來源，不清會留著舊內容
        for r in range(SKEL_START + DETAIL_SHIFT,
                       max(SKEL_ROWS, n_skel) + DETAIL_SHIFT + 1):
            for c in range(1, SKEL_COLS + 1):
                put(ws, r, c, None)
        for (r, c), v in sorted(tpl2.items(), reverse=True):
            nr = r + (DETAIL_SHIFT if r > FROZEN_ROWS else 0)
            if (nr, c) in fresh_detail:      # 新科目由 fresh_detail 直接寫最終公式
                continue
            new, ok, unresolved = rewrite(v, (map_a, map_h, DETAIL_SHIFT, 0),
                                          rowmap_detail, blockmap_detail, new_blocks,
                                          f"'{TOTAL_SHEET}'!")
            new, ok = polish(v, new, ok, c, nr)
            if not ok:                       # 對不到新位置，只能清空
                failed.append((name, nr, c, v, why(unresolved, blockmap_detail)))
                new = None
            put(ws, nr, c, new)
        for (nr, c), v in fresh_detail.items():
            put(ws, nr, c, v)
        for c in range(1, SKEL_COLS + 4):                 # 借總表的欄位說明填新空列
            put(ws, FROZEN_ROWS + 1, c, wb[TOTAL_SHEET].cell(FROZEN_ROWS + 1, c).value)
        put(ws, 1, 12, idx)                               # L1 = 校序號
        put(ws, 3, 1, f"中華民國{year}年度")
        put(ws, 5, 2, f"{year - 1}.12.31")
        put(ws, 5, 7, f"{year}.12.31")
        put(ws, 7, 12, f"{year}年12月")
        last = n_skel + DETAIL_SHIFT
        if ws.max_row > last:
            ws.delete_rows(last + 1, ws.max_row - last + 1)
        base = name.split("-")[0]
        if base != name:
            renamed.append((name, base))
            ws.title = base

    # ---- 總表：骨架重排 + 改寫公式 + 寫入資料區塊 ----
    # 總表的骨架自成一份（它多一列欄位說明，所以整體比明細樣板低 DETAIL_SHIFT 列），
    # 先換算回樣板座標套同一組對照表，寫回去時再加回來
    wt = wb[TOTAL_SHEET]
    tpl_t, dropped_total = reposition(
        {(r - DETAIL_SHIFT, c): wt.cell(r, c).value
         for r in range(1, SKEL_ROWS + DETAIL_SHIFT + 1)
         for c in range(1, SKEL_COLS + 4)},
        map_a, map_h)
    fresh_total = fresh_cells("", DETAIL_SHIFT)
    for r in range(SKEL_START + DETAIL_SHIFT,
                   max(SKEL_ROWS, n_skel) + DETAIL_SHIFT + 1):
        for c in range(1, SKEL_COLS + 4):
            put(wt, r, c, None)
    for (r, c), v in sorted(tpl_t.items(), reverse=True):
        nr = r + DETAIL_SHIFT
        if (nr, c) in fresh_total:
            continue
        new, ok, unresolved = rewrite(v, (map_a, map_h, DETAIL_SHIFT, DETAIL_SHIFT),
                                      rowmap_total, blockmap_total, new_blocks, "")
        new, ok = polish(v, new, ok, c, nr)
        if not ok:
            failed.append((TOTAL_SHEET, nr, c, v, why(unresolved, blockmap_total)))
            new = None
        put(wt, nr, c, new)
    for (r, c), v in fresh_total.items():
        put(wt, r, c, v)
    put(wt, 3, 1, f"中華民國{year}年度")
    put(wt, 5, 2, f"{year - 1}.12.31")
    put(wt, 5, 7, f"{year}.12.31")
    first, last = wb.sheetnames[2], wb.sheetnames[-1]
    realigned = set()
    for r in range(1, n_skel + DETAIL_SHIFT + 1):
        for c in range(1, SKEL_COLS + 1):
            v = wt.cell(r, c).value
            if isinstance(v, str) and ":" in v and "!" in v and v.startswith("=SUM("):
                v = re.sub(r"[^(]+!", f"{first}:{last}!", v)
                # 一律指向同一列，才能整欄往下拉；範本原有幾格是跳列的
                fixed = re.sub(r"!(\$?[A-Z]{1,2}\$?)\d+", lambda m: f"!{m.group(1)}{r}", v)
                if fixed != re.sub(r"!(\$?[A-Z]{1,2}\$?)(\d+)",
                                   lambda m: f"!{m.group(1)}{int(m.group(2)) + DETAIL_SHIFT}", v):
                    realigned.add(r)
                put(wt, r, c, fixed)

    tail = n_skel + DETAIL_SHIFT
    if wt.max_row > tail:
        wt.delete_rows(tail + 1, wt.max_row - tail + 1)
    for name in BLOCK_NAMES:
        b = new_blocks[name]
        head = "項      目" if name == "現流" else "科目"
        write_block(wt, b.header - 1, b.header, markers[name], head,
                    unis_cur, new_data[name], num_fmt[name])

    # ---- 報告 ----
    report(f"資料區塊（總表）：")
    for name in BLOCK_NAMES:
        b = new_blocks[name]
        report(f"  {name:<6} 表頭 {b.header:>4}　資料 {b.start}–{b.end}（{len(new_data[name])} 列）")

    report(f"\n骨架（{SKEL_START}–{n_skel} 列，範本 {SKEL_ROWS} 列）："
           f"平衡表科目 {len(plan_a)} 項（縮排 ≤ {MAX_LEVEL} 層）、現流項目 {len(plan_h)} 項")
    for side, src, units, plan, gone in (
            ("平衡表", bal_labels, units_a, plan_a, gone_a),
            ("現流", cf_labels, units_h, plan_h, gone_h)):
        added = [parented(src)[j] for j, _, u in plan if u is None]
        old_shown = parented([lb for lb, _ in units])
        if added:
            report(f"  {side}新增 {len(added)} 列（機械公式已補，調整欄留白待填）：")
            for name in added:
                report(f"    + {name}")
        if gone:
            report(f"  {side}移除 {len(gone)} 列（今年來源沒有這一項）：")
            for i in gone:
                report(f"    − {old_shown[i]}")
    for side, drop in (("明細表", dropped_detail), ("總表", dropped_total)):
        if drop:
            cells = "、".join(sorted({f"{get_column_letter(c)}{r}" for r, c, _ in drop},
                                    key=lambda s: (s[0], int(s[1:]))))
            report(f"  {side}隨移除列一起丟棄的儲存格 {len(drop)} 格：{cells}")

    if realigned:
        report(f"\n總表跨表加總已改為一律指同一列（範本原本跳列的第 "
              f"{'、'.join(str(r) for r in sorted(realigned))} 列已修正）")

    if renamed:
        report(f"\n表名已去除去年的註記後綴：")
        for old, new in renamed:
            report(f"  {old}  →  {new}")
    if failed:
        WHY = [
            ("ref",     "範本裡本來就是 #REF! 斷鏈"),
            ("void",    "指向骨架上沒有科目名的空列，範本原本就是空參照"),
            ("removed", "指向今年來源沒有、已被移除的科目／項目"),
            ("stale",   "指向明細表已移除的舊鏡射區、新版沒有對應位置"),
            ("block",   "指向資料區塊、但對不到新位置"),
            ("lost",    "! 對不到位置也歸不了類，可能是邊界判斷有誤"),
        ]
        buckets = collections.defaultdict(list)
        for f in failed:
            buckets["ref" if "#REF!" in str(f[3]) else f[4]].append(f)

        def cells(items):
            return "、".join(sorted({f"{get_column_letter(c)}{r}" for _, r, c, *_ in items},
                                   key=lambda s: (s[0], int(s[1:]))))

        report(f"\n已清空 {len(failed)} 個公式（座標為產出檔）：")
        for key, text in WHY:
            got = buckets.get(key)
            if not got:
                continue
            report(f"  {text}：{cells(got)}（{len(got)} 格）")
            if key in ("block", "lost"):
                for name, r, c, v, _ in got[:5]:
                    report(f"      {name} {get_column_letter(c)}{r}: {v}")

    DRIFT_SKIP = range(14, SKEL_COLS + 1)          # N~R：檢查欄，各表擺放位置本就不一
    ties = [(r, c, mc) for r, c, mc in drift
            if c not in DRIFT_SKIP and c != 12 and mc[0][1] < N_UNI * 0.75]
    n_check = sum(1 for r, c, _ in drift if c in DRIFT_SKIP)
    if ties:
        report(f"\n樣板取眾數時的近平手（{len(ties)} 處，建議人工確認）：")
        for r, c, mc in ties:
            opts = "　/　".join(f"{str(v)[:32]}×{n}" for v, n in mc)
            report(f"  {get_column_letter(c)}{r}: {opts}")
    report(f"（另有 N~R 檢查欄 {n_check} 格各表擺法不一，已統一為眾數）")

    out = args.output or str(
        pathlib.Path(outdir)
        / f"大學{year}現金流量表勾稽檔-{N_UNI}所大學{'-空白' if blank else ''}.xlsx")
    wb.save(out)
    report(f"\n✓ 已寫出 {pathlib.Path(out).name}")
    return out, "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description="產生大學現金流量表勾稽底稿")
    ap.add_argument("balance", help="12.平衡xxxx.xlsx")
    ap.add_argument("cashflow", help="3.現流xxxx.xlsx")
    ap.add_argument("-t", "--template", default=DEFAULT_TEMPLATE)
    ap.add_argument("-o", "--output")
    ap.add_argument("--year", type=int, help="覆寫解析到的年度")
    ap.add_argument("--bal-sheet"), ap.add_argument("--prev-sheet"), ap.add_argument("--cf-sheet")
    ap.add_argument("--blank", action="store_true",
                    help="空白版：清掉範本累積的借貸代號與調整公式，只留科目與機械公式")
    ap.add_argument("--strip-template", metavar="OUT",
                    help="把範本瘦身成只留骨架的 OUT（另存 OUT.mirror.json），不做其他事")
    args = ap.parse_args()
    if args.strip_template:
        strip_template(args.template, args.strip_template)
        return
    try:
        _, text = build(args.balance, args.cashflow, args.template, args.output,
                        args.year, args.bal_sheet, args.prev_sheet, args.cf_sheet,
                        blank=args.blank)
    except BuildError as e:
        raise SystemExit(f"✗ {e}")
    print(text)


if __name__ == "__main__":
    main()
